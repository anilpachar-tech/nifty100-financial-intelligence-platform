from openpyxl import load_workbook

wb = load_workbook("output/screener_output.xlsx")

print(wb.sheetnames)