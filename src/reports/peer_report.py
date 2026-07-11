"""
Sprint 3 - Day 20

Peer Comparison Excel Report
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side

DB_PATH = "db/nifty100.db"

OUTPUT_FILE = "output/peer_comparison.xlsx"


def load_data():

    conn = sqlite3.connect(DB_PATH)

    companies = pd.read_sql(
        "SELECT * FROM companies",
        conn
    )

    financial = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    peers = pd.read_sql(
        "SELECT * FROM peer_groups",
        conn
    )

    percentiles = pd.read_sql(
        "SELECT * FROM peer_percentiles",
        conn
    )

    conn.close()

    return (
        companies,
        financial,
        peers,
        percentiles
    )

def prepare_master():

    companies, financial, peers, percentiles = load_data()

    master = (
        financial
        .merge(
            companies[
                [
                    "id",
                    "company_name"
                ]
            ],
            left_on="company_id",
            right_on="id",
            how="left",
            suffixes=("", "_company")
        )
        .merge(
            peers[
                [
                    "company_id",
                    "peer_group_name",
                    "is_benchmark"
                ]
            ],
            on="company_id",
            how="left"
        )
    )

    return master, percentiles

def prepare_percentiles(percentiles):

    percentile_wide = (
        percentiles
        .pivot_table(
            index=[
                "company_id",
                "year"
            ],
            columns="metric",
            values="percentile_rank",
            aggfunc="first"
        )
        .reset_index()
    )

    percentile_wide.columns.name = None

    return percentile_wide

def prepare_report_data(master, percentile_wide):

    report = master.merge(
        percentile_wide,
        on=[
            "company_id",
            "year"
        ],
        how="left",
        suffixes=("", "_percentile")
    )

    return report

def create_workbook():

    wb = Workbook()

    wb.remove(
        wb.active
    )

    return wb


def create_peer_sheets(wb, report):

    peer_groups = (
        report[
            "peer_group_name"
        ]
        .dropna()
        .unique()
    )

    peer_groups = sorted(peer_groups)

    for group in peer_groups:

        wb.create_sheet(
            title=group[:31]
        )

    return wb

# ============================================================
# Report Configuration
# ============================================================

METRIC_COLUMNS = [

    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "operating_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "asset_turnover",
    "free_cash_flow_cr",
    "capex_cr",
    "earnings_per_share",
    "book_value_per_share",
    "dividend_payout_ratio_pct",
    "total_debt_cr",
    "cash_from_operations_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "eps_cagr_5yr",
    "revenue_cagr_3yr",
    "composite_quality_score",
    "dividend_yield"

]

PERCENTILE_COLUMNS = [

    "return_on_equity_pct_percentile",
    "return_on_capital_employed_pct_percentile",
    "net_profit_margin_pct_percentile",
    "debt_to_equity_percentile",
    "interest_coverage_percentile",
    "asset_turnover_percentile",
    "free_cash_flow_cr_percentile",
    "pat_cagr_5yr_percentile",
    "revenue_cagr_5yr_percentile",
    "eps_cagr_5yr_percentile"

]


def latest_company_data(report):

    latest = (

        report
        .sort_values("year")
        .groupby("company_id", as_index=False)
        .tail(1)

    )

    return latest

def write_sheet_data(wb, latest):

    headers = [

        "company_id",
        "company_name",
        "is_benchmark"
    ] + METRIC_COLUMNS + PERCENTILE_COLUMNS


    for sheet in wb.sheetnames:

        ws = wb[sheet]

        ws.append(headers)

        peer_df = latest[
            latest["peer_group_name"] == sheet
        ]

        for row in peer_df.itertuples(index=False):

            values = [

                row.company_id,
                row.company_name,
                row.is_benchmark

            ]

            for col in METRIC_COLUMNS:

                values.append(
                    getattr(
                        row,
                        col,
                        None
                    )
                )

            for col in PERCENTILE_COLUMNS:

                values.append(
                    getattr(
                        row,
                        col,
                        None
                    )
                )

            ws.append(values)

    return wb

def format_workbook(wb):

    header_fill = PatternFill(
        fill_type="solid",
        start_color="1F4E78",
        end_color="1F4E78"
    )

    header_font = Font(
        bold=True,
        color="FFFFFF"
    )

    gold_fill = PatternFill(
        fill_type="solid",
        start_color="FFD966",
        end_color="FFD966"
    )

    green_fill = PatternFill(
        fill_type="solid",
        start_color="C6EFCE",
        end_color="C6EFCE"
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        start_color="FFEB9C",
        end_color="FFEB9C"
    )

    red_fill = PatternFill(
        fill_type="solid",
        start_color="FFC7CE",
        end_color="FFC7CE"
    )

    thin = Side(style="thin")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    percentile_start = 3 + len(METRIC_COLUMNS) + 1

    for ws in wb.worksheets:

        # Header
        for cell in ws[1]:

            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center"
            )

        # Benchmark Row
        benchmark_column = None

        for row in range(2, ws.max_row + 1):

            company = ws.cell(
                row=row,
                column=3
            ).value

            benchmark = ws.cell(
                row=row,
                column=4
            ).value

            if benchmark == "1":

                for cell in ws[row]:

                    cell.fill = gold_fill
                    cell.font = Font(bold=True)

            # benchmark check later
            # currently skip if company missing

        # Percentile Colors
        for row in range(2, ws.max_row + 1):

            for col in range(
                percentile_start,
                ws.max_column + 1
            ):

                cell = ws.cell(
                    row=row,
                    column=col
                )

                if cell.value is None:
                    continue

                try:

                    value = float(cell.value)

                except:

                    continue

                if value >= 75:

                    cell.fill = green_fill

                elif value <= 25:

                    cell.fill = red_fill

                else:

                    cell.fill = yellow_fill

                cell.border = border

        # Auto Width
        for column_cells in ws.columns:

            length = max(

                len(str(cell.value))
                if cell.value is not None
                else 0

                for cell in column_cells

            )

            ws.column_dimensions[
                column_cells[0].column_letter
            ].width = min(
                length + 3,
                25
            )

        median_row = ws.max_row + 1

        ws.cell(
            row=median_row,
            column=1
        ).value = "Median"

        start_metric = 4

        end_metric = 3 + len(METRIC_COLUMNS)

        for col in range(start_metric, end_metric + 1):

            values = []

            for r in range(2, ws.max_row + 1):

                value = ws.cell(
                    row=r,
                    column=col
                ).value

                if isinstance(value, (int, float)):

                    values.append(value)
            if values:

                ws.cell(
                    row=median_row,
                    column=col
                ).value = round(
                    pd.Series(values).median(),
                    2
                )
        for cell in ws[median_row]:

            cell.font = Font(
                bold=True
            )

        # Benchmark Row (Apply Last)

        for row in range(2, ws.max_row):
            
            benchmark = ws.cell(
                row=row,
                column=3
            ).value

            if benchmark == "1":

                for cell in ws[row]:

                    cell.fill = gold_fill
                    cell.font = Font(bold=True)

    return wb

if __name__ == "__main__":

    companies, financial, peers, percentiles = load_data()

    master, percentiles = prepare_master()

    percentile_wide = prepare_percentiles(
        percentiles
    )

    report = prepare_report_data(
        master,
        percentile_wide
    )

    wb = create_workbook()

    print("=" * 60)
    print("Database Loaded")
    print("=" * 60)

    print("Companies :", len(companies))
    print("Financial Ratios :", len(financial))
    print("Peer Groups :", len(peers))
    print("Peer Percentiles :", len(percentiles))

    print()
    wb = create_peer_sheets(wb, report)

    latest = latest_company_data(report)

    wb = write_sheet_data(wb, latest)

    wb = format_workbook(wb)

    print("=" * 60)
    print("Master Dataset")
    print("=" * 60)

    print("Rows :", len(master))
    print()

    print(
        master[
            [
                "company_id",
                "company_name",
                "year",
                "peer_group_name",
                "is_benchmark"
            ]
        ].head()
    )

    print()

    print("=" * 60)
    print("Peer Percentiles (wide)")
    print("=" * 60)

    print("Rows :", len(percentile_wide))
    print()

    print(
        percentile_wide.head()
    )

    print()

    print("=" * 60)
    print("Final Report Dataset")
    print("=" * 60)

    print("Rows :", len(report))
    print()

    print(
        report[
            [
                "company_id",
                "company_name",
                "peer_group_name",
                "return_on_equity_pct",
                "return_on_equity_pct_percentile",
                "debt_to_equity",
                "debt_to_equity_percentile"
            ]
        ].head()
    )

    print()

    print("=" * 60)
    print("Workbook Created")
    print("=" * 60)

    print("Total Sheets :", len(wb.sheetnames))
    print()

    for sheet in wb.sheetnames:
        print(f"Sheet: {sheet}")

    print()

    print("=" * 60)
    print("Peer Sheets Populated")
    print("=" * 60)

    for sheet in wb.sheetnames:

        ws = wb[sheet]

        print(f"Sheet: {sheet} - Rows: {ws.max_row}")

    print()

    print("=" * 60)
    print("Workbook Formatted")
    print("=" * 60)

    print("Header formatting applied")

    print("Conditional formatting applied")

    print("Column widths adjusted")

    wb.save(
        OUTPUT_FILE
    )

    print()

    print("=" * 60)
    print("Workbook Saved")
    print("=" * 60)
    
    print(OUTPUT_FILE)