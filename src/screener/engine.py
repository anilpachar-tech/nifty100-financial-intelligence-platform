"""
Sprint 3 - Day 15

Financial Screener Engine
"""

import sqlite3
from warnings import filters
import pandas as pd
import yaml
import os
from src.screener.composite_score import calculate_composite_score
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

DB_PATH = "db/nifty100.db"
CONFIG_PATH = "config/screener_config.yaml"


def load_data():

    conn = sqlite3.connect(DB_PATH)

    ratios = pd.read_sql(
        "SELECT * FROM financial_ratios",
        conn
    )

    sectors = pd.read_sql(
        "SELECT company_id, broad_sector FROM sectors",
        conn
    )

    conn.close()

    return ratios, sectors


def load_config():

    with open(
        CONFIG_PATH,
        "r",
        encoding="utf-8"
    ) as file:

        return yaml.safe_load(file)

def prepare_master_dataset():

    ratios, sectors = load_data()

    conn = sqlite3.connect(DB_PATH)

    profit = pd.read_sql(
        """
        SELECT
            company_id,
            year,
            sales
        FROM profitandloss
        """,
        conn
    )

    conn.close()

    master = (
        ratios
        .merge(
            sectors,
            on="company_id",
            how="left"
        )
        .merge(
            profit,
            on=["company_id", "year"],
            how="left"
        )
    )

    master = (
        master
        .sort_values("year")
        .groupby("company_id", as_index=False)
        .tail(1)
    )

    master = master.reset_index(drop=True)

    return master

def apply_filters(
    df: pd.DataFrame,
    filters: dict,
    skip_financial_de: bool = True
) -> pd.DataFrame:

    result = df.copy()

    # ROE Filter
    if "roe_min" in filters:
        result = result[
            result["return_on_equity_pct"] >=
            filters["roe_min"]
        ]

    # Debt to Equity Filter
    if "debt_to_equity_max" in filters:

        if skip_financial_de:

            financial_mask = (
                result["broad_sector"] 
                .fillna("")
                .str.lower()
                == "financials" 
            )

            financial = result[financial_mask]

            non_financial = result[~financial_mask]

            non_financial = non_financial[
                non_financial["debt_to_equity"]
                <= filters["debt_to_equity_max"]
            ]

            result = pd.concat(
                [financial, non_financial],
                ignore_index=True
            )
        else:

            result = result[
                result["debt_to_equity"]
                <= filters["debt_to_equity_max"]
            ]
            
    # Free Cash Flow
    if "free_cash_flow_min" in filters:
        result = result[
            result["free_cash_flow_cr"] >=
            filters["free_cash_flow_min"]
        ]

    # Revenue CAGR
    if "revenue_cagr_5yr_min" in filters:
        result = result[
            result["revenue_cagr_5yr"] >=
            filters["revenue_cagr_5yr_min"]
        ]

    # Revenue CAGR 3 Year
    if "revenue_cagr_3yr_min" in filters:
        result = result[
            result["revenue_cagr_3yr"] >=
            filters["revenue_cagr_3yr_min"]
        ]

    # Debt Declining
    if filters.get("debt_declining"):
        result = result[
            result["debt_declining"] == 1
        ]

    # PAT CAGR
    if "pat_cagr_5yr_min" in filters:
        result = result[
            result["pat_cagr_5yr"] >=
            filters["pat_cagr_5yr_min"]
        ]
    # Operating Profit Margin
    if "opm_min" in filters:
        result = result[
            result["operating_profit_margin_pct"] >=
            filters["opm_min"]
        ]

    # Interest Coverage
    if "interest_coverage_min" in filters:
        result = result[
            result["interest_coverage"] >=
            filters["interest_coverage_min"]
        ]

    # Asset Turnover
    if "asset_turnover_min" in filters:
        result = result[
            result["asset_turnover"] >=
            filters["asset_turnover_min"]
        ]

    # Sales
    if "sales_min" in filters:

        if "sales" in result.columns:

            result = result[
                result["sales"] >=
                filters["sales_min"]
            ]
    # P/E
    if "pe_max" in filters:
        result = result[
            result["pe"] <= filters["pe_max"]
        ]

    # P/B
    if "pb_max" in filters:
        result = result[
            result["pb"] <= filters["pb_max"]
        ]

    # Dividend Yield
    if "dividend_yield_min" in filters:
        result = result[
            result["dividend_yield"] >=
            filters["dividend_yield_min"]
        ]
    return result

if __name__ == "__main__":

    master = prepare_master_dataset()

    master = calculate_composite_score(master)

    config = load_config()

    print("=" * 60)
    print("Master Dataset")
    print("=" * 60)

    print("Rows :", len(master))
    print()

    print(master[
        [
            "company_id",
            "year",
            "sales",
            "return_on_equity_pct",
            "debt_to_equity",
            "composite_score",
            "sector_relative_score",
            "broad_sector"
        ]
    ].head())

    print()
    print("=" * 60)
    print("Quality Compounder Screener")
    print("=" * 60)

    quality = apply_filters(
        master,
        config["quality_compounder"]
    )

    quality = quality.sort_values(
    "composite_score",
    ascending=False
    )

    print("Companies Found :", len(quality))
    print()

    print(
        quality[
            [
                "company_id",
                "year",
                "sales",
                "return_on_equity_pct",
                "debt_to_equity",
                "free_cash_flow_cr",
                "revenue_cagr_5yr"
            ]
        ].head(20)
    )

    print()
    print("=" * 60)
    print("Growth Accelerator Screener")
    print("=" * 60)

    growth = apply_filters(
        master,
        config["growth_accelerator"]
    )

    growth = growth.sort_values(
        "composite_score",
        ascending=False
    )

    print("Companies Found :", len(growth))
    print()

    print(
        growth[
            [
                "company_id",
                "year",
                "sales",
                "pat_cagr_5yr",
                "revenue_cagr_5yr",
                "debt_to_equity",
                "composite_score"
            ]
        ].head(20)
    )
    
    print()
    print("=" * 60)
    print("Value Pick Screener")
    print("=" * 60)

    value = apply_filters(
        master,
        config["value_pick"]
    )

    print("Companies Found :", len(value))
    print()

    print(
        value[
            [
                "company_id",
                "year",
                "sales",
                "pe",
                "pb",
                "dividend_yield",
                "debt_to_equity",
                "composite_score"
            ]
        ].head(20)
    )

    print()
    print("=" * 60)
    print("Turnaround Watch Screener")
    print("=" * 60)

    turnaround = apply_filters(
        master,
        config["turnaround_watch"]
    )

    turnaround = turnaround.sort_values(
        "composite_score",
        ascending=False
    )

    print("Companies Found :", len(turnaround))
    print()

    print(
        turnaround[
            [
                "company_id",
                "year",
                "revenue_cagr_3yr",
                "debt_declining",
                "return_on_equity_pct",
                "composite_score"
            ]
        ].head(20)
    )

    print()
    print("=" * 60)
    print("Dividend Champion Screener")
    print("=" * 60)

    dividend = apply_filters(
        master,
        config["dividend_champion"]
    )

    dividend = dividend.sort_values(
        "composite_score",
        ascending=False
    )

    print("Companies Found :", len(dividend))
    print()

    print(
        dividend[
            [
                "company_id",
                "year",
                "sales",
                "dividend_yield",
                "dividend_payout_ratio_pct",
                "free_cash_flow_cr",
                "composite_score"
            ]
        ].head(20)
    )

    print()
    print("=" * 60)
    print("Debt-Free Blue Chip Screener")
    print("=" * 60)

    bluechip = apply_filters(
    master,
    config["debt_free_blue_chip"],
    skip_financial_de=False
    )

    bluechip = bluechip.sort_values(
        "composite_score",
        ascending=False
    )

    print("Companies Found :", len(bluechip))
    print()

    print(
    bluechip[
        [
            "company_id",
            "year",
            "sales",
            "return_on_equity_pct",
            "debt_to_equity",
            "composite_score"
        ]
    ].head(20)
    
)
print()
print("=" * 60)
print("Generating Excel Report")
print("=" * 60)

os.makedirs("output", exist_ok=True)

green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

red_fill = PatternFill(
    start_color="FFC7CE",
    end_color="FFC7CE",
    fill_type="solid"
)

header_fill = PatternFill(
    start_color="4F81BD",
    end_color="4F81BD",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    color="FFFFFF"
)

wb = Workbook()

def format_sheet(ws):

    # Header Formatting
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Auto Width
    for column in ws.columns:

        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        ws.column_dimensions[column_letter].width = max_length + 2

    # Cell Colors
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2):

        for cell in row:

            column = headers[cell.column - 1]

            if not isinstance(cell.value, (int, float)):
                continue

            # Higher is Better
            if column in [
                "return_on_equity_pct",
                "return_on_capital_employed_pct",
                "free_cash_flow_cr",
                "revenue_cagr_5yr",
                "pat_cagr_5yr",
                "composite_score",
                "interest_coverage",
                "sector_relative_score"
            ]:

                if cell.value >= 15:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

            # Lower is Better
            elif column in [
                "debt_to_equity",
                "pe",
                "pb"
            ]:

                if cell.value <= 1:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill


# ---------- Quality Compounder ----------
ws = wb.active
ws.title = "Quality Compounder"

for row in quality.itertuples(index=False):
    if ws.max_row == 1:
        ws.append(list(quality.columns))
    ws.append(list(row))

format_sheet(ws)


# ---------- Growth Accelerator ----------
ws = wb.create_sheet("Growth Accelerator")

for row in growth.itertuples(index=False):
    if ws.max_row == 1:
        ws.append(list(growth.columns))
    ws.append(list(row))

format_sheet(ws)

# ---------- Value Pick ----------
ws = wb.create_sheet("Value Pick")

for row in value.itertuples(index=False):
    if ws.max_row == 1:
        ws.append(list(value.columns))
    ws.append(list(row))

format_sheet(ws)

# ---------- Turnaround Watch ----------
ws = wb.create_sheet("Turnaround Watch")

for row in turnaround.itertuples(index=False):
    if ws.max_row == 1:
        ws.append(list(turnaround.columns))
    ws.append(list(row))

format_sheet(ws)

# ---------- Dividend Champion ----------
ws = wb.create_sheet("Dividend Champion")

for row in dividend.itertuples(index=False):
    if ws.max_row == 1:
        ws.append(list(dividend.columns))
    ws.append(list(row))

format_sheet(ws)

# ---------- Debt-Free Blue Chip ----------
ws = wb.create_sheet("Debt Free Blue Chip")

for row in bluechip.itertuples(index=False):
    if ws.max_row == 1:
        ws.append(list(bluechip.columns))
    ws.append(list(row))

format_sheet(ws)

output_file = "output/screener_output.xlsx"

wb.save(output_file)

print("Excel Generated Successfully")
print(output_file)